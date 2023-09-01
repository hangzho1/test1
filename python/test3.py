# n,td,tc = map(int,input().split())
# damage = list1(map(int,input().split()))
# coin = list1(map(int,input().split()))
# count = 0
# def dfs(damage,begin,coin,count,td,tc):
#     if tc <=0:
#         return 
#     if  td < 0:
#         count += 1
#         return 
#     for index in range(begin,n):
#         dfs(damage,index + 1,coin,count,td-damage[index],tc-coin[index])

# dfs(damage,0,coin,count,td,tc)
# print(count)a




# def permute(nums):
#     """
#     :type nums: list1[int]
#     :rtype: list1[list1[int]]
#     """
#     ### size是数组的大小，depth是当前遍历的深度，它的大小和path大小一样
#     ##path是存储已遍历点的数组，used是用来判断原数组哪些点已经用了，还是否可用，res是返回最后的结果
#     def dfs(nums, size, depth, path, used, res):
#             if depth == size:
#                 ##这里着实不懂为什么不是path,而是path[:]呢
#                 res.append(path[:])
#                 return

#             for i in range(size):
#                 if not used[i]:
#                     used[i] = True
#                     path  += [nums[i]]
#                     # path.append(nums[i])

#                     dfs(nums, size, depth+1, path, used, res)

#                     used[i] = False
#                     path.pop()

#     size = len(nums)
#     if len(nums) == 0:
#         return []

#     used = [False for _ in range(size)]
#     res = []
#     dfs(nums, size, 0, [], used, res)
#     return res
# print(permute([1,2,3]))

# nums = list1(map(int,input().split()))
# target = int(input())


# nums = [2,3,6,7]
# target = 7
# def combinationSum(target):
#     def dfs(nums,count,begin,path,target):
#         if target < 0:
#             return 
#         elif target == 0:
#             count +=1
#             res.append(path[:])
#             return count
#         for i in range(begin,len(nums)):
#             path.append(nums[i])
#             dfs(nums,count,i,path,target - nums[i])
#             path.pop()
#         return count
#     res = []
#     path = []
#     if min(nums) > target:
#         return res

#     count = dfs(nums,0,0,path,target)
#     return res,count
# print(combinationSum(target))




    
# 360 烽火台
# n,m,x,k = 5,5,1,2
# nums = [4,4,2,4,4]
""" n,m,x,k = 5,5,1,2
nums = [4,4,2,4,4]
def combinationSum(nums):
    def jia(nums,i,x,k):
        for c in range(i-x,i+x+1):
            if 0 <= c < len(nums):
                nums[c] +=k
        return min(nums)
    def dfs(nums,depth,maxCount,path):
        if depth == m:
            res.append(path) ###输出路径
            count.append(maxCount) ##把每一种可能的结果算出来
            return 
        for i in range(len(nums)):
            nums1 = nums.copy() ## 这里的复制其实就是复制上一次
            maxCount = jia(nums1,i,x,k)
            dfs(nums1,depth + 1,maxCount,path + [i])
            nums1 = nums.copy()
        return count
    maxCount = 0
    res,path,count = [], [],[]
    dfs(nums,0,0,path)
    return count
    # return res,count
print(combinationSum(nums)) """

"""
思迈特组数之和 
n, target = map(int,input().split())
nums = [x for x in range(1,n+1)]
def dfs(nums,begin,path,res,target):
    if target == 0:
        res.append(path)
        return 
    for i in range(begin,n):
        if target < 0:
            break
        dfs(nums,i +1,path + [nums[i]],res,target - nums[i])

path,res = [],[]
dfs(nums,0,path,res,target)

for i in res:
    print(*i)
 """






# n,td,tc = map(int,input().split())
# damage = list1(map(int,input().split()))
# coin = list1(map(int,input().split()))
# pathDamage,list1Damage,list1Coin,pathCoin = [],[],[],[]
# def fightMonster(damage,td,tc):
#     if sum(damage) <= td:
#         return 0 
#     def dfs(damage,coin,list1Damage,list1Coin,pathDamage,pathCoin,begin,td,tc):
#         if tc <=0:
#             return 
#         if  td < 0:
#             list1Coin.append(sum(pathCoin))
#             list1Damage.append(pathDamage)
#             if begin < len(damage):
#                 dfs(damage,coin,list1Damage,list1Coin,pathDamage+[damage[begin]],pathCoin+ [coin[begin]],begin,td-damage[begin],tc-coin[begin])
#             return 
#         for index in range(begin,n):
#             dfs(damage,coin,list1Damage,list1Coin,pathDamage+[damage[index]],pathCoin+ [coin[index]],index+1,td-damage[index],tc-coin[index])

#     dfs(damage,coin,list1Damage,list1Coin,pathDamage,pathCoin,0,td,tc)
# fightMonster(damage,td,tc)
# print(pathDamage,list1Damage,list1Coin,pathCoin)

# n,td,tc = map(int,input().split())
# damage = list1(map(int,input().split()))
# coin = list1(map(int,input().split()))
# pathDamage,list1Damage,list1Coin,pathCoin = [],[],[],[]
# def fightMonster(td,tc):
#     if sum(damage) <= td:
#         return 0 
#     def dfs(pathDamage,pathCoin,begin,td,tc):
#         if tc <=0:
#             return 
#         if  td < 0:
#             list1Coin.append(sum(pathCoin))
#             list1Damage.append(pathDamage)
#             if begin < len(damage):
#                 dfs(pathDamage+[damage[begin]],pathCoin+ [coin[begin]],begin,td-damage[begin],tc-coin[begin])
#             return 
#         for index in range(begin,n):
#             dfs(pathDamage+[damage[index]],pathCoin+ [coin[index]],index+1,td-damage[index],tc-coin[index])

#     dfs(pathDamage,pathCoin,0,td,tc)
# fightMonster(td,tc)
# print(pathDamage,list1Damage,list1Coin,pathCoin)



























# import random
# nums = [random.randint(-10,101) for x in range(10)]
# print(nums)
# for i in range(len(nums) -1):
#     flag = True
#     for j in range(len(nums) - i -1):
#         if(nums[j+1] < nums[j]):
#             nums[j+1],nums[j] = nums[j],nums[j+1]
#             flag = False
#     if flag == True:
#         break
# print(nums)

# s = "hello"
# nums = list1(s)
# hashmap = {"a","e","i","o","u"}
# left,right =0,len(nums)-1
# while left < right:
#     while left <right and  nums[right] not in hashmap:
#         right -=1
#     while left < right and nums[left] not in hashmap:
#         left +=1
#     nums[left],nums[right] = nums[right],nums[left]
#     left+=1
#     right-=1
# print("".join(nums))


# s = "Let's take LeetCode contest"
# s = ["".join(reversed(x)) for x in s.split()]
# print(" ".join(s))










# import random
# nums = [random.randint(1,100) for _ in range(10)]
# print(nums)
# left,right = 0,len(nums)  -1
# while left < right:
#     while left < right and nums[right] % 2 == 0 :
#         right -=1
#     while left < right and nums[left] % 2 != 0:
#         left += 1 
#     nums[left],nums[right] = nums[right],nums[left]
# print(nums)






# n,l,r = map(int,input().split())
# nums = []
# nums.append(n)
# def dfs(nums):
#     flag = True
#     for i in range(len(nums)):
#         if(nums[i] > 1):
#             temp = nums.pop(i)
#             nums.insert(i,temp// 2)
#             nums.insert(i+1,temp % 2)
#             nums.insert(i+2,temp // 2)
#             flag = False
#     if flag:
#         return 
#     else:
#         dfs(nums)
# dfs(nums)
# print(sum(nums[l-1:r])

# nums = [2,1,2]
# print(sum(nums[1:3]))

# string = "abcdesAssayEaaassYyy"
# res,count = 0,0
# key_word = "easy"
# for s in string:
#     if s.lower() == key_word[count]:
#         count += 1
#     if count == 4:
#         count = 0
#         res += 1
# print(res)



# m,n  = map(int,input())
# list1 = [[] for _ in range(m)]

# list1[0][1] = 1
# print(list1)
# Scanner sc  = new Scanner(System.in);
# int m = sc.nextInt(), n = sc.nextInt();
# sc.nextLine();
# for (int i = 0; i < list1.length; i++) {
#     char[] a = sc.nextLine().toCharArray();
#     for (int j = 0; j < a.length; j++) {
#         list1[i][j] = a[j];
#     }
# }
#         System.out.println(redFlower(list1));

#     }
#     public static int redFlower(char[][] list1){
#         int redflower = 0,m = list1.length,n = list1[0].length;
#         for (int i = 0; i < list1.length; i++) {
#             for (int j = 0; j < list1[0].length; j++) {
# //                if(list1[i][i] == 'M'){
# //                    continue;
# //                }
#                   if(list1[i][j] == 'F'){
#                     if(i -1 >= 0 && list1[i-1][j] == 'F'){
#                         redflower++;
#                     }
#                     if(i -1 >= 0 && j+ 1 < n && list1[i-1][j+1] == 'F'){
#                         redflower++;
#                     }
#                     if(j + 1 < n && list1[i][j+1] == 'F'){
#                         redflower++;
#                     }
#                     if(i +1 < m && j+ 1 < n && list1[i + 1][j + 1] == 'F'){
#                         redflower++;
#                     }
#                     if(i +1 < m   && list1[i+1][j] == 'F'){
#                         redflower++;
#                     }
#                     if(i + 1 < m && j-1 >= 0 && list1[i+1][j-1] == 'F'){
#                         redflower++;
#                     }
#                     if(j -1 >= 0 && list1[i][j-1] == 'F'){
#                         redflower++;
#                     }
#                     if(i -1 >= 0  && j- 1 >=0 && list1[i-1][j-1] == 'F'){
#                         redflower++;
#                     }
#                 }
#             }
#         }
#         return redflower/2;
#     }
# }





# nums = [ [1,2], [2,3], [3,4], [1,3] ]
# nums.sort(key = lambda x:x[1])
# print(nums)
# end = -float("inf")
# cnt = 0
# for i in nums:
#     start = i[0]
#     if start >= end:
#         cnt +=1
#         end = i[1]
# print(len(nums) - cnt)

# from PIL import Image,ImageOps


# def resize_pic(in_name,size):
#     img = Image.open(in_name)
#     img  = ImageOps.fit(img,(34000,size),Image.ANTIALIAS)
#     return img
# image  = "E:\Desktop/微信图片_2022010414413817.jpg"
# print(Image.open(image).size)
# backgroud =  Image.new("RGB",Image.open(image).size,(255,255,255))
# backgroud.save("E:/Desktop/back.jpg")
# img = resize_pic(image,20000)
# img.show()
# img.save("E:/Desktop/test2-1.jpg")









# import os
# import sys
# import requests
# cloud_config_api='http://10.67.119.15:8899/cloud/api/cloud_config/'

# #这个portal.py的用处倒底是啥呀，只跟vaas相关联吗，还是jenkins里面的某些操作也需要用到？
# def get_cloud_config(url):
#     ## get cloud config from portal.api
#     _cloud_config = []
#     proxies = {
#         'http': None,
#         'https': None
#     }
#     response = requests.get(url=url, proxies=proxies)
#     response.raise_for_status()
#     if response.status_code == requests.codes.ok:
#         response_json_list = response.json()
#         if len(response_json_list) > 0:
#             #print("Get config successfully")
#             # print(response_json_list)
#             _cloud_config = response_json_list
#         else:
#             print("No valid config")
#     else:
#         print("Stop to get config due to {}".format(response.text))

#     return _cloud_config


# def get_cloud_exclude_list():
#     exclude_list = []
#     cloud_config = get_cloud_config(url='http://10.67.119.15:8899/cloud/api/setting/')
#     print(cloud_config)
#     for i in cloud_config:
#         if i['name'] == 'exclude_workload_list':
#             exclude_list = i['value'].split(',')

#     return exclude_list# get_cloud_config(cloud_config_api)

# print(get_cloud_exclude_list())












# import os
# import re
# import sys

# def get_platforms(workload):
#     platform_list = []
#     work_dir = "/home/hangz/hangzinfra"
#     if os.path.isdir(os.path.join(work_dir, "workload", workload)):
#         # '/home/hangz/hangzinfra/workload/Nginx/CMakeLists.txt'
#         cmake_file_path = os.path.join(work_dir, "workload", workload, "CMakeLists.txt")
#         if os.path.isfile(cmake_file_path):
#             # '/home/hangz/hangzinfra/workload/Nginx/cmake'
#             wl_cmake_dir = os.path.join(work_dir,  "workload", workload, "cmake")
#             if os.path.isdir(wl_cmake_dir):
#                 # 列表列出所有这个目录下所有的文件 ['ARMv9.cmake', 'SPR.cmake', 'MILAN.cmake', 'ICX.cmake', 'common.cmake', 'GNR.cmake', 'ARMv8.cmake']
#                 pf_cmakes = os.listdir(wl_cmake_dir)
#                 if "PR_" not in work_dir:
#                     if 'SPR.cmake' in pf_cmakes and 'SPR' not in platform_list:
#                         platform_list.append('SPR')
#                     if "ICX.cmake" in pf_cmakes:
#                         platform_list.append('ICX')
#                     if "MILAN.cmake" in pf_cmakes:
#                         platform_list.append('MILAN')
#                     if 'MILAN.cmake' not in pf_cmakes and 'ROME.cmake' in pf_cmakes:
#                         platform_list.append('ROME')
#                     if 'ARMv9.cmake' in pf_cmakes:
#                         platform_list.append('ARMv9')
#                     if 'ARMv9.cmake' not in pf_cmakes and 'ARMv8.cmake' in pf_cmakes:
#                         platform_list.append('ARMv8')
#                 else:
#                     if "SPR.cmake" in pf_cmakes:
#                         platform_list.append('SPR')

#                     # New cmake model, Construct the AMD platform list for Cloud
#                     elif "MILAN.cmake" in pf_cmakes or "ROME.cmake" in pf_cmakes:
#                         if 'ROME.cmake' in pf_cmakes and 'MILAN.cmake' in pf_cmakes:
#                             platform_list.append('MILAN')
#                         elif 'ROME.cmake' not in pf_cmakes and 'MILAN.cmake' in pf_cmakes:
#                             platform_list.append('MILAN')
#                         elif 'ROME.cmake' in pf_cmakes and 'MILAN.cmake' not in pf_cmakes:
#                             platform_list.append('ROME')

#                     # New cmake model, Construct the ARM platform list for Cloud
#                     elif "ARMv9.cmake" in pf_cmakes or "ARMv8.cmake" in pf_cmakes:
#                         if 'ARMv8.cmake' in pf_cmakes and 'ARMv9.cmake' in pf_cmakes:
#                             platform_list.append('ARMv9')
#                         elif 'ARMv8.cmake' not in pf_cmakes and 'ARMv9.cmake' in pf_cmakes:
#                             platform_list.append('ARMv9')
#                         elif 'ARMv8.cmake' in pf_cmakes and 'ARMv9.cmake' not in pf_cmakes:
#                             platform_list.append('ARMv8')
#                             # ['SPR', 'ICX', 'MILAN', 'ARMv9']
#                 platform_list = list(set(platform_list))
#                 print(platform_list)
#             else:
#                 with open(cmake_file_path, 'r', encoding="utf8") as fp:
#                     content = fp.read()
#                     pt_list = re.findall(r'if.*\(\"(.*)\" MATCHES', content)

#                     if len(pt_list) <= 0 :
#                         pt_list = re.findall(r'if.*\(PLATFORM STREQUAL \"(.*)\"', content)

#                     if len(pt_list) > 0 :
#                         for pt in pt_list:
#                             if 'SPR' in pt and 'SPR' not in platform_list:
#                                 platform_list.append('SPR')
#                             if "PR_" not in work_dir:
#                                 if 'ICX' in pt:
#                                     platform_list.append('ICX')
#                                 if 'MILAN' in pt:
#                                     platform_list.append('MILAN')
#                                 if 'MILAN' not in pt and 'ROME' in pt:
#                                     platform_list.append('ROME')
#                                 if 'ARMv9' in pt:
#                                     platform_list.append('ARMv9')
#                                 if 'ARMv9' not in pt and 'ARMv8' in pt:
#                                     platform_list.append('ARMv8')
#                             else:
#                                 if 'ICX' in pt and 'SPR' not in platform_list:
#                                     platform_list.append('SPR')
#                                 # Construct the AMD platform list for Cloud
#                                 elif 'MILAN' in pt or 'ROME' in pt:
#                                     if 'ROME' in pt and 'MILAN' in pt:
#                                         platform_list.append('MILAN')
#                                     elif 'ROME' not in pt and 'MILAN' in pt:
#                                         platform_list.append('MILAN')
#                                     elif 'ROME' in pt and 'MILAN' not in pt:
#                                         platform_list.append('ROME')

#                                 # Construct the ARM platform list for Cloud
#                                 elif 'ARMv9' in pt or 'ARMv8' in pt:
#                                     if 'ARMv8' in pt and 'ARMv9' in pt:
#                                         platform_list.append('ARMv9')
#                                     elif 'ARMv8' not in pt and 'ARMv9' in pt:
#                                         platform_list.append('ARMv9')
#                                     elif 'ARMv8' in pt and 'ARMv9' not in pt:
#                                         platform_list.append('ARMv8')
#                                 else:
#                                     if 'SPR' in platform_list:
#                                         continue
#                                     platform_list.append(pt.split(" ")[1])

#                         platform_list = list(set(platform_list))
#                         print(platform_list)
#                     else:
#                         print(f"No supported platform")
#     return platform_list



# print(get_platforms("Nginx"))







# a="'--set AWS_WORKER_CPU_MODEL_REGEX='8375' --set AWS_WORKER_INSTANCE_TYPE=m6i.4xlarge --set AWS_MACHINE_TYPE=m6i.4xlarge --set AWS_ZONE=us-east-2a --set AWS_REGION=us-east-2 --set AWS_CLIENT_INSTANCE_TYPE=m6i.4xlarge"
# print(a.strip("'"))
# print(a)



import os
import re
import sys
import openpyxl


class Infra_Dynamic(object):
    def __init__(self, sheet):
        self.sheet = sheet
        self.file_path = "/home/hangz/hangz/python/Infra_dynamic.xlsx"


    def _csv_filter_workload_name(self, testcase):
        workload_name = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            filter_case_cell = "{}{}".format('C', row)
            if work_sheet[wl_name_cell].value == testcase :
                workload_name = work_sheet[filter_case_cell].value
                break
        return workload_name
    
    def _csv_filter_csp_options(self, testcase):
        csp_type = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            filter_case_cell = "{}{}".format('G', row)
            if work_sheet[wl_name_cell].value == testcase :
                csp_type = work_sheet[filter_case_cell].value
                break
        return csp_type
    
    def _csv_filter_case(self, testcase):
        filter_case = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            filter_case_cell = "{}{}".format('J', row)
            if work_sheet[wl_name_cell].value == testcase :
                filter_case = work_sheet[filter_case_cell].value
                break
        return filter_case

    def _csv_filter_terraform_options(self, testcase):
        terraform_options = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            terraform_options_cell = "{}{}".format('H', row)
            if work_sheet[wl_name_cell].value == testcase :
                terraform_options = work_sheet[terraform_options_cell].value.strip("'") if work_sheet[terraform_options_cell].value else work_sheet[terraform_options_cell].value
                break
        return terraform_options

    def _csv_filter_ctest_options(self, testcase):
        ctest_options = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            ctest_options_cell = "{}{}".format('I', row)
            if work_sheet[wl_name_cell].value == testcase :
                ctest_options = work_sheet[ctest_options_cell].value.strip("'") if work_sheet[ctest_options_cell].value else work_sheet[ctest_options_cell].value
                break
        return ctest_options

    def _csv_exclude_case(self, testcase):
        exclude_case = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            exclude_case_cell = "{}{}".format('K', row)
            if work_sheet[wl_name_cell].value == testcase :
                exclude_case = work_sheet[exclude_case_cell].value
                break
        return exclude_case

    def _csv_parallel_run_case_num(self, testcase):
        _parallel_run_case = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            parallel_case_num_cell = "{}{}".format('L', row)
            if work_sheet[wl_name_cell].value == testcase :
                _parallel_run_case = work_sheet[parallel_case_num_cell].value
                break
        return _parallel_run_case

    def _csv_workload_params(self, testcase):
        _workload_params = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            workload_params_cell = "{}{}".format('M', row)
            if work_sheet[wl_name_cell].value == testcase :
                _workload_params = work_sheet[workload_params_cell].value
                break
        return _workload_params

    def _csv_workload_cluster_file(self, testcase):
        _workload_cluster_file = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            workload_cluster_file_cell = "{}{}".format('N', row)
            if work_sheet[wl_name_cell].value == testcase :
                _workload_cluster_file = work_sheet[workload_cluster_file_cell].value
                break
        return _workload_cluster_file
    
    def _csv_workload_platform_case(self, testcase):
        platform_name = None
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        for row in range(2, work_sheet.max_row + 1):
            wl_name_cell = "{}{}".format('B', row)
            workload_cluster_file_cell = "{}{}".format('O', row)
            if work_sheet[wl_name_cell].value == testcase :
                platform_name = work_sheet[workload_cluster_file_cell].value
                break
        return platform_name

    def _csv_filter_testcase(self,workload_list):
        """
        file_path: the full path of special_workload.xlsx
        """
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        cloud_coverd_list = []
        static_coverd_list = []
        related_coverd_list = []
        case_colum = 'B'
        column = 'C'  
        include_colum = 'D'   
        bm_colum = 'E'  
        cloud_colum = 'F'  
        for row in range(2, work_sheet.max_row + 1):
            case_name = "{}{}".format(case_colum, row)
            cell_name = "{}{}".format(column, row)
            include_column_name = "{}{}".format(include_colum, row)
            bm_column_name = "{}{}".format(bm_colum, row)
            cloud_column_name = "{}{}".format(cloud_colum, row)
            if work_sheet[cell_name].value not in workload_list or (work_sheet[include_column_name].value and work_sheet[include_column_name]) == "no":
                continue
            elif work_sheet[cell_name].value == "Nginx" and work_sheet[bm_column_name].value and work_sheet[bm_column_name].value.lower() == "yes":
                static_coverd_list.append(work_sheet[case_name].value)
            elif work_sheet[cell_name].value == "Nginx" and work_sheet[cloud_column_name] and work_sheet[cloud_column_name].value.lower() == "yes":
                cloud_coverd_list.append(work_sheet[case_name].value)
            else:
                related_coverd_list.append(work_sheet[case_name].value)
        return cloud_coverd_list, static_coverd_list, related_coverd_list

    def filter_workload_name(self, testcase):
        return self._csv_filter_workload_name(testcase)
    
    def filter_workload_csp_options(self, testcase):
        return self._csv_filter_csp_options(testcase)
    
    def filter_workload_terraform_options(self, testcase):
        return self._csv_filter_terraform_options(testcase)
    
    def filter_workload_ctest_options(self, testcase):
        return self._csv_filter_ctest_options(testcase)
    
    def filter_workload_case(self, testcase):
        return self._csv_filter_case(testcase)

    def exclude_workload_case(self, testcase):
        return self._csv_exclude_case(testcase)

    def parallel_case_num(self, testcase):
        return self._csv_parallel_run_case_num(testcase)

    def workload_params(self, testcase):
        return self._csv_workload_params(testcase)

    def workload_cluster_file(self, testcase):
        return self._csv_workload_cluster_file(testcase)
    
    def workload_platform_name(self, testcase):
        return self._csv_workload_platform_case(testcase)

    def filter_group_workloads(self):
        """
        Return all workload list which belong to 'sheet' group.
        """
        group_workload_list = []
        wb = openpyxl.load_workbook(self.file_path)
        work_sheet= wb[self.sheet]
        column = 'C'  ## Column C is 'workload'
        include_colum = 'D'   ## Column D is 'include?'
        for row in range(2, work_sheet.max_row + 1):
            cell_name = "{}{}".format(column, row)
            include_column_name = "{}{}".format(include_colum, row)
            if work_sheet[include_column_name].value and work_sheet[include_column_name].value.lower() == 'yes' and work_sheet[cell_name].value not in group_workload_list:
                group_workload_list.append(work_sheet[cell_name].value)
        group_workload_list = list(set(group_workload_list) - set(['']))
        # print("test_workload_list {}".format(group_workload_list), len(group_workload_list))
        return group_workload_list

    def filter_tested_workloads(self, workload_list):
        cloud_tested_list = []
        static_tested_list = []
        related_tested_list = []
        cloud_tested_list, static_tested_list, related_tested_list = self._csv_filter_testcase(workload_list)
        print("static %s" %static_tested_list,len(static_tested_list))
        print("cloud %s" %cloud_tested_list, len(cloud_tested_list))
        print("related %s" %related_tested_list,len(related_tested_list))
        return cloud_tested_list, static_tested_list, related_tested_list


if __name__ == '__main__':

    func = sys.argv[1]
    testname = sys.argv[2]
    # workload_list = testname.strip('][').split(', ') 
    workload_list = [element.strip() for element in testname.strip('][').split(',')]
    workload_list = list(set(workload_list) - set(['']))
    sheet = sys.argv[3]
    dynamic = Infra_Dynamic(sheet)
    if func == 'sheet':
        print(dynamic.filter_group_workloads())
    if func == 'workload':
        print(dynamic.filter_tested_workloads(workload_list))
    if func == 'case':
        print(dynamic.filter_workload_name(testname), dynamic.filter_workload_csp_options(testname), dynamic.filter_workload_terraform_options(testname), dynamic.filter_workload_ctest_options(testname), 
              dynamic.filter_workload_case(testname), dynamic.exclude_workload_case(testname), dynamic.parallel_case_num(testname), dynamic.workload_params(testname), dynamic.workload_cluster_file(testname),dynamic.workload_platform_name(testname))


